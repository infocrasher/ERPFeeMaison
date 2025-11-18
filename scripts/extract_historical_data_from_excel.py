#!/usr/bin/env python3
"""
Script d'extraction des données historiques depuis les fichiers Excel de comptabilité.

Ce script :
1. Parcourt tous les fichiers Excel dans le dossier
2. Extrait les données de la feuille RECAP (CA journalier, charges, salaires, loyer)
3. Extrait les données de la feuille Charges (achats avec consolidation)
4. Génère un CSV structuré pour l'import dans Prophet

Usage:
    python scripts/extract_historical_data_from_excel.py "Téléchargements Comptabilité" output.csv
"""

import sys
import os
import csv
import re
from datetime import datetime, date
from decimal import Decimal
from collections import defaultdict
import openpyxl

def find_recap_sheet(workbook):
    """Trouve la feuille RECAP (gère les variations de nom)"""
    for sheet_name in workbook.sheetnames:
        if 'recap' in sheet_name.lower() or 'récap' in sheet_name.lower():
            return workbook[sheet_name]
    return None

def find_charges_sheet(workbook):
    """Trouve la feuille Charges (gère les variations de nom)"""
    for sheet_name in workbook.sheetnames:
        if 'charge' in sheet_name.lower() and 'salair' not in sheet_name.lower():
            return workbook[sheet_name]
    return None

def parse_date(cell_value):
    """Parse une date depuis une cellule Excel"""
    if cell_value is None:
        return None
    
    if isinstance(cell_value, datetime):
        return cell_value.date()
    
    if isinstance(cell_value, date):
        return cell_value
    
    # Essayer de parser une chaîne
    if isinstance(cell_value, str):
        # Formats possibles : "2025-01-01", "01/01/2025", etc.
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
            try:
                return datetime.strptime(cell_value.strip(), fmt).date()
            except:
                continue
    
    return None

def parse_amount(value):
    """Parse un montant (gère les virgules, espaces, texte)"""
    if value is None:
        return Decimal('0.0')
    
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    
    if isinstance(value, str):
        # Nettoyer : enlever espaces, "DA", virgules de milliers
        cleaned = value.strip().replace(' ', '').replace(',', '').replace('DA', '').replace('da', '').replace('DZD', '').replace('dzd', '')
        # Remplacer virgule décimale par point
        cleaned = cleaned.replace(',', '.')
        try:
            return Decimal(cleaned)
        except:
            return Decimal('0.0')
    
    return Decimal('0.0')

def extract_recap_data(sheet, filename):
    """Extrait les données de la feuille RECAP"""
    data = []
    
    # Trouver la ligne d'en-têtes
    header_row = None
    date_col = None
    revenue_col = None
    charges_col = None
    salary_col = None
    rent_col = None
    
    # Chercher les en-têtes (lignes 1-3)
    for row_idx in range(1, min(4, sheet.max_row + 1)):
        row = sheet[row_idx]
        for col_idx, cell in enumerate(row, 1):
            value = str(cell.value).lower() if cell.value else ''
            if 'date' in value and date_col is None:
                date_col = col_idx
            # Recette : chercher "recette" mais ignorer les colonnes de comparaison mensuelle
            # Prendre la colonne "Recette " (avec espace) ou "Recette" sans nom de mois
            if ('recette' in value or 'revenu' in value or 'ca' in value) and revenue_col is None:
                # Ignorer les colonnes avec des noms de mois (janvier, février, mars, etc.)
                months_in_value = any(month in value for month in [
                    'janvier', 'février', 'fevrier', 'mars', 'avril', 'mai', 'juin',
                    'juillet', 'août', 'aout', 'septembre', 'octobre', 'novembre', 'décembre', 'decembre'
                ])
                # Ignorer aussi les colonnes avec des années (2024, 2025, etc.)
                has_year = any(str(year) in value for year in range(2019, 2026))
                
                if not months_in_value and not has_year:
                    revenue_col = col_idx
            if 'charge' in value and charges_col is None:
                charges_col = col_idx
            if 'salair' in value and salary_col is None:
                salary_col = col_idx
            if 'loyer' in value and rent_col is None:
                rent_col = col_idx
    
    if not date_col:
        print(f"⚠️  Colonne Date non trouvée dans {filename}")
        return data
    
    # Extraire les totaux mensuels depuis les lignes de résumé (sans date, en bas)
    monthly_totals = {
        'revenue': Decimal('0.0'),
        'charges': Decimal('0.0'),
        'salary': Decimal('0.0'),
        'rent': Decimal('0.0')
    }
    
    # Chercher les totaux mensuels dans toutes les lignes
    # Les totaux sont généralement des valeurs élevées (> 100k pour recette/charges, > 100k pour salaires, 40k-200k pour loyer)
    # et peuvent être dans des lignes avec ou sans date valide
    
    for row_idx in range(2, sheet.max_row + 1):
        date_cell = sheet.cell(row_idx, date_col)
        date_value = parse_date(date_cell.value)
        
        # Chercher les valeurs élevées qui pourraient être des totaux mensuels
        if revenue_col:
            rev_val = parse_amount(sheet.cell(row_idx, revenue_col).value)
            # Totaux mensuels de recette sont généralement > 500,000
            if rev_val > 500000:
                monthly_totals['revenue'] = max(monthly_totals['revenue'], rev_val)
        
        if charges_col:
            chg_val = parse_amount(sheet.cell(row_idx, charges_col).value)
            # Totaux mensuels de charges sont généralement > 500,000
            if chg_val > 500000:
                monthly_totals['charges'] = max(monthly_totals['charges'], chg_val)
        
        if salary_col:
            sal_val = parse_amount(sheet.cell(row_idx, salary_col).value)
            # Totaux mensuels de salaires sont généralement > 300,000
            if sal_val > 300000:
                monthly_totals['salary'] = max(monthly_totals['salary'], sal_val)
        
        if rent_col:
            rent_val = parse_amount(sheet.cell(row_idx, rent_col).value)
            # Loyer entre 40k et 200k
            if 40000 <= rent_val <= 200000:
                monthly_totals['rent'] = max(monthly_totals['rent'], rent_val)
    
    monthly_rent = monthly_totals['rent']
    monthly_salary = monthly_totals['salary']
    
    # monthly_salary est déjà extrait dans monthly_totals ci-dessus
    
    # Parcourir les lignes de données
    for row_idx in range(2, sheet.max_row + 1):
        row = sheet[row_idx]
        
        # Extraire la date
        date_cell = sheet.cell(row_idx, date_col)
        record_date = parse_date(date_cell.value)
        
        if not record_date:
            continue
        
        # Extraire les montants journaliers
        revenue = Decimal('0.0')
        charges = Decimal('0.0')
        salary = Decimal('0.0')
        rent = Decimal('0.0')
        
        if revenue_col:
            revenue = parse_amount(sheet.cell(row_idx, revenue_col).value)
            # Si valeur très élevée (> 500k) ET pas de date valide, c'est probablement un total mensuel
            # On l'ignore pour cette ligne (elle sera traitée dans les totaux mensuels)
            # MAIS si on a une date valide, c'est peut-être une valeur journalière réelle (jour exceptionnel)
            # On garde donc la valeur si on a une date valide
        
        if charges_col:
            charges = parse_amount(sheet.cell(row_idx, charges_col).value)
            # Même logique : garder si date valide, ignorer si pas de date
        
        if salary_col:
            salary = parse_amount(sheet.cell(row_idx, salary_col).value)
            # Même logique : garder si date valide, ignorer si pas de date
        
        if rent_col:
            rent = parse_amount(sheet.cell(row_idx, rent_col).value)
            # Même logique : garder si date valide, ignorer si pas de date
        
        # IMPORTANT: Utiliser les valeurs journalières réelles telles quelles
        # Ne répartir les totaux mensuels QUE si les valeurs journalières sont absentes (0 ou vide)
        # ET seulement pour les lignes avec une date valide (pas les lignes de résumé)
        if record_date:
            import calendar
            days_in_month = calendar.monthrange(record_date.year, record_date.month)[1]
            
            # Recette : utiliser la valeur journalière réelle si disponible (> 0)
            # Sinon, répartir le total mensuel uniquement pour les jours sans données
            if revenue == 0 and monthly_totals['revenue'] > 0:
                # Pas de valeur journalière, répartir le total mensuel
                revenue = monthly_totals['revenue'] / Decimal(str(days_in_month))
            # Sinon, garder la valeur journalière réelle (même si > 500k, c'est une valeur réelle)
            
            # Charges : utiliser la valeur journalière si disponible, sinon répartir le total mensuel
            # Les charges sont généralement mensuelles, donc on répartit toujours si total disponible
            if monthly_totals['charges'] > 0:
                # Toujours utiliser le total mensuel réparti pour les charges (car généralement mensuelles)
                charges = monthly_totals['charges'] / Decimal(str(days_in_month))
            elif charges > 500000:
                # Si valeur très élevée sans total mensuel, c'est probablement un total, on l'ignore
                charges = Decimal('0.0')
            
            # Salaires : utiliser la valeur journalière si disponible, sinon répartir le total mensuel
            # Les salaires sont généralement mensuels, donc on répartit toujours si total disponible
            if monthly_totals['salary'] > 0:
                # Toujours utiliser le total mensuel réparti pour les salaires (car généralement mensuels)
                salary = monthly_totals['salary'] / Decimal(str(days_in_month))
            elif salary > 300000:
                # Si valeur très élevée sans total mensuel, c'est probablement un total, on l'ignore
                salary = Decimal('0.0')
            
            # Loyer : utiliser la valeur journalière si disponible, sinon répartir le total mensuel
            # Le loyer est généralement mensuel, donc on répartit toujours si total disponible
            if monthly_rent > 0:
                # Toujours utiliser le total mensuel réparti pour le loyer (car généralement mensuel)
                rent = monthly_rent / Decimal(str(days_in_month))
            elif rent > 200000:
                # Si valeur très élevée sans total mensuel, c'est probablement un total, on l'ignore
                rent = Decimal('0.0')
        
        # Ignorer les lignes sans données significatives
        if revenue == 0 and charges == 0 and salary == 0 and rent == 0:
            continue
        
        data.append({
            'date': record_date,
            'revenue': revenue,
            'charges': charges,
            'salary': salary,
            'rent': rent
        })
    
    return data

def extract_charges_data(sheet, filename):
    """Extrait et consolide les données de la feuille Charges"""
    # Dictionnaire pour consolider : {nom_produit: {total_qty, total_value, count}}
    products = defaultdict(lambda: {'total_qty': Decimal('0.0'), 'total_value': Decimal('0.0'), 'count': 0})
    
    # Trouver la ligne d'en-têtes
    header_row = None
    product_col = None
    qty_col = None
    price_col = None
    total_col = None
    
    # Chercher les en-têtes (lignes 1-3)
    for row_idx in range(1, min(4, sheet.max_row + 1)):
        row = sheet[row_idx]
        for col_idx, cell in enumerate(row, 1):
            value = str(cell.value).lower() if cell.value else ''
            if 'produit' in value and product_col is None:
                product_col = col_idx
            if 'quantité' in value or 'qty' in value and qty_col is None:
                qty_col = col_idx
            if ('prix' in value and 'unitaire' in value) or 'unit' in value and price_col is None:
                price_col = col_idx
            if ('prix' in value and 'total' in value) or ('prix' in value and price_col and col_idx != price_col) and total_col is None:
                total_col = col_idx
    
    if not product_col:
        print(f"⚠️  Colonne Produit non trouvée dans Charges de {filename}")
        return products
    
    # Parcourir les lignes de données
    for row_idx in range(2, sheet.max_row + 1):
        product_cell = sheet.cell(row_idx, product_col)
        product_name = str(product_cell.value).strip() if product_cell.value else ''
        
        if not product_name or len(product_name) < 2:
            continue
        
        # Extraire quantité et prix
        qty = Decimal('0.0')
        price = Decimal('0.0')
        total = Decimal('0.0')
        
        if qty_col:
            qty = parse_amount(sheet.cell(row_idx, qty_col).value)
        
        if price_col:
            price = parse_amount(sheet.cell(row_idx, price_col).value)
        
        if total_col:
            total = parse_amount(sheet.cell(row_idx, total_col).value)
        
        # Si on a le total mais pas le prix unitaire, le calculer
        if total > 0 and qty > 0 and price == 0:
            price = total / qty
        
        # Si on a le prix unitaire mais pas le total, le calculer
        if price > 0 and qty > 0 and total == 0:
            total = price * qty
        
        if qty > 0 or total > 0:
            # Normaliser le nom du produit (enlever espaces multiples, mettre en minuscules pour regroupement)
            product_name_normalized = re.sub(r'\s+', ' ', product_name.strip().lower())
            
            products[product_name_normalized]['total_qty'] += qty
            products[product_name_normalized]['total_value'] += total
            products[product_name_normalized]['count'] += 1
    
    return products

def parse_filename_date(filename):
    """Extrait le mois et l'année depuis le nom du fichier"""
    # Patterns possibles : "Mai 2019", "Janvier 2025", "Mars2020", etc.
    patterns = [
        r'(\w+)\s+(\d{4})',  # "Mai 2019"
        r'(\w+)(\d{4})',      # "Mars2020"
        r'(\d{4})',           # Juste l'année
    ]
    
    months_fr = {
        'janvier': 1, 'février': 2, 'fevrier': 2, 'mars': 3, 'avril': 4, 'mai': 5, 'juin': 6,
        'juillet': 7, 'août': 8, 'aout': 8, 'septembre': 9, 'octobre': 10, 'novembre': 11, 'décembre': 12, 'decembre': 12
    }
    
    for pattern in patterns:
        match = re.search(pattern, filename, re.IGNORECASE)
        if match:
            if len(match.groups()) == 2:
                month_str = match.group(1).lower()
                year = int(match.group(2))
                month = months_fr.get(month_str, 1)
                return date(year, month, 1)
            elif len(match.groups()) == 1:
                year = int(match.group(1))
                return date(year, 1, 1)
    
    return None

def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/extract_historical_data_from_excel.py <dossier_excel> [output.csv]")
        print("\nExemple:")
        print("  python scripts/extract_historical_data_from_excel.py 'Téléchargements Comptabilité' donnees_historiques.csv")
        sys.exit(1)
    
    input_dir = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else 'donnees_historiques_comptabilite.csv'
    
    if not os.path.isdir(input_dir):
        print(f"❌ Dossier non trouvé: {input_dir}")
        sys.exit(1)
    
    # Lister tous les fichiers Excel
    excel_files = [f for f in os.listdir(input_dir) if f.endswith(('.xlsx', '.xls'))]
    excel_files.sort()
    
    print(f"📂 {len(excel_files)} fichiers Excel trouvés")
    print(f"📄 Extraction vers: {output_file}\n")
    
    # Dictionnaire pour agréger par date
    daily_data = defaultdict(lambda: {
        'revenue': Decimal('0.0'),
        'charges': Decimal('0.0'),
        'salary': Decimal('0.0'),
        'rent': Decimal('0.0')
    })
    
    # Dictionnaire pour consolider les achats par mois
    monthly_purchases = defaultdict(lambda: {
        'total_value': Decimal('0.0'),
        'products': defaultdict(lambda: {'total_qty': Decimal('0.0'), 'total_value': Decimal('0.0')})
    })
    
    processed = 0
    errors = 0
    
    for filename in excel_files:
        filepath = os.path.join(input_dir, filename)
        print(f"📄 Traitement: {filename}")
        
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            
            # Extraire données RECAP
            recap_sheet = find_recap_sheet(wb)
            if recap_sheet:
                recap_data = extract_recap_data(recap_sheet, filename)
                for record in recap_data:
                    d = record['date']
                    daily_data[d]['revenue'] += record['revenue']
                    daily_data[d]['charges'] += record['charges']
                    daily_data[d]['salary'] += record['salary']
                    daily_data[d]['rent'] += record['rent']
                print(f"   ✅ RECAP: {len(recap_data)} jours extraits")
            else:
                print(f"   ⚠️  Feuille RECAP non trouvée")
            
            # Extraire données Charges
            charges_sheet = find_charges_sheet(wb)
            if charges_sheet:
                charges_data = extract_charges_data(charges_sheet, filename)
                # Déterminer le mois depuis le nom du fichier
                file_date = parse_filename_date(filename)
                if file_date:
                    month_key = f"{file_date.year}-{file_date.month:02d}"
                    for product, data in charges_data.items():
                        monthly_purchases[month_key]['total_value'] += data['total_value']
                        monthly_purchases[month_key]['products'][product]['total_qty'] += data['total_qty']
                        monthly_purchases[month_key]['products'][product]['total_value'] += data['total_value']
                    print(f"   ✅ Charges: {len(charges_data)} produits consolidés")
                else:
                    print(f"   ⚠️  Impossible de déterminer le mois depuis le nom du fichier")
            
            processed += 1
            
        except Exception as e:
            print(f"   ❌ Erreur: {e}")
            errors += 1
            import traceback
            traceback.print_exc()
            continue
    
    print(f"\n📊 Résumé:")
    print(f"   Fichiers traités: {processed}")
    print(f"   Erreurs: {errors}")
    print(f"   Jours avec données: {len(daily_data)}")
    print(f"   Période: {min(daily_data.keys()) if daily_data else 'N/A'} à {max(daily_data.keys()) if daily_data else 'N/A'}")
    
    # Écrire le CSV
    print(f"\n💾 Écriture du CSV: {output_file}")
    
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['date', 'revenue', 'purchases', 'salaries', 'rent', 'other_expenses'])
        
        # Trier par date
        for d in sorted(daily_data.keys()):
            data = daily_data[d]
            # Les charges viennent déjà des totaux mensuels répartis dans extract_recap_data
            # On utilise directement data['charges'] qui est déjà le total mensuel réparti par jour
            purchases = data['charges']
            
            writer.writerow([
                d.isoformat(),
                float(data['revenue']),
                float(purchases),
                float(data['salary']),
                float(data['rent']),
                0.0  # other_expenses (peut être enrichi plus tard)
            ])
    
    print(f"✅ CSV créé avec succès: {output_file}")
    print(f"   {len(daily_data)} enregistrements")
    
    # Créer aussi un résumé des achats consolidés
    summary_file = output_file.replace('.csv', '_achats_consolides.csv')
    print(f"\n💾 Écriture du résumé des achats: {summary_file}")
    
    with open(summary_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['mois', 'produit', 'quantite_totale', 'valeur_totale', 'prix_moyen_pondere'])
        
        for month_key in sorted(monthly_purchases.keys()):
            month_data = monthly_purchases[month_key]
            for product, product_data in sorted(month_data['products'].items()):
                if product_data['total_qty'] > 0:
                    avg_price = product_data['total_value'] / product_data['total_qty']
                    writer.writerow([
                        month_key,
                        product,
                        float(product_data['total_qty']),
                        float(product_data['total_value']),
                        float(avg_price)
                    ])
    
    print(f"✅ Résumé des achats créé: {summary_file}")

if __name__ == '__main__':
    main()

